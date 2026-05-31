import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .validation import is_url


def generate_workbook(tables, out_file):
    logging.info(f"generating workbook {out_file}")
    wb = Workbook()
    sheet = wb.active

    row = 1

    for table in tables:
        logging.debug(f"adding table {table['title']} to workbook")
        sheet.merge_cells(
            f'A{row}:{chr(len(table["headers"]) - 1 + ord("A"))}{row}'
        )
        header_cell = sheet[f'A{row}']
        header_cell.value = table['title']
        header_cell.font = Font(bold=True, size=11, color='FFFFFF')
        header_cell.fill = PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid'
        )
        header_cell.alignment = Alignment(horizontal='center')
        row += 1

        for col, col_name in enumerate(table['headers'], 1):
            cell = sheet.cell(row=row, column=col, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'
            )
        row += 1

        for data_row in table['data']:
            for col, val in enumerate(data_row, 1):
                cell = sheet.cell(row=row, column=col, value=val)

                if isinstance(val, str) and is_url(val):
                    display_val = val[:47] + "..." if len(val) > 50 else val
                    cell.value = display_val
                    cell.hyperlink = val
                    cell.style = "Hyperlink"

            row += 1

        row += 2

    logging.debug(f"saving workbook to {out_file}")
    wb.save(out_file)
